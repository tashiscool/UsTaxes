#!/usr/bin/env python3
"""Build a machine-readable 1040-family workbook parity harness.

This script inspects the local forecast workbooks, groups sheets into the
1040-family forms that UsTaxes implements, and emits both an inventory and a
comparison matrix under docs/.
"""

from __future__ import annotations

import argparse
import csv
import json
import warnings
from dataclasses import dataclass, asdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Sequence

from openpyxl import load_workbook


ROOT = Path("/Users/tkhan/IdeaProjects/taxes/UsTaxes")
DEFAULT_WORKBOOKS = [
    Path("/Users/tkhan/Downloads/24_1040.xlsx"),
    Path("/Users/tkhan/Downloads/25_1040.xlsx"),
    Path("/Users/tkhan/Downloads/26_1040 Forecast (Rev. 41).xlsx"),
]
DEFAULT_OUTPUT_DIR = ROOT / "docs"


@dataclass(frozen=True)
class FormGroupSpec:
    canonical_form: str
    primary_sheets: tuple[str, ...]
    support_sheets: tuple[str, ...]
    local_files: tuple[str, ...]
    local_test_files: tuple[str, ...]
    notes: str


@dataclass
class SheetInventoryRow:
    workbook_label: str
    workbook_path: str
    sheet_name: str
    canonical_form: str | None
    role: str
    cell_count: int
    formula_count: int
    formula_density: float


@dataclass
class WorkbookSummary:
    workbook_label: str
    workbook_path: str
    sheet_count: int
    target_sheet_count: int
    support_sheet_count: int
    excluded_sheet_count: int
    total_cells: int
    total_formulas: int
    sheets: list[SheetInventoryRow]


@dataclass
class FormCoverageRow:
    canonical_form: str
    local_files: list[str]
    local_test_files: list[str]
    local_file_count: int
    local_present: bool
    workbook_2024_present: bool
    workbook_2024_sheet_names: list[str]
    workbook_2024_cell_count: int
    workbook_2024_formula_count: int
    workbook_2025_present: bool
    workbook_2025_sheet_names: list[str]
    workbook_2025_cell_count: int
    workbook_2025_formula_count: int
    workbook_2026_present: bool
    workbook_2026_sheet_names: list[str]
    workbook_2026_cell_count: int
    workbook_2026_formula_count: int
    years_present: list[str]
    years_missing: list[str]
    notes: str


FORM_GROUPS: tuple[FormGroupSpec, ...] = (
    FormGroupSpec(
        canonical_form="W-2",
        primary_sheets=("W-2s",),
        support_sheets=(),
        local_files=("FormW2W3.ts",),
        local_test_files=(),
        notes="Workbook wage intake sheet and the local wage statement implementation."
    ),
    FormGroupSpec(
        canonical_form="1099s",
        primary_sheets=(
            "1099-INT",
            "1099-DIV",
            "1099-G",
            "1099-MISC",
            "1099-NEC",
            "1099-R",
        ),
        support_sheets=("SSA-1099", "Lines 4ab (IRA Dist.)"),
        local_files=(
            "Form1099.ts",
            "F1099INT.ts",
            "F1099DIV.ts",
            "F1099G.ts",
            "F1099MISC.ts",
            "F1099NEC.ts",
            "F1099R.ts",
        ),
        local_test_files=(),
        notes="Aggregates the information-return bundle the workbooks use for 1099 and SSA/IRA-adjacent inputs."
    ),
    FormGroupSpec(
        canonical_form="1040",
        primary_sheets=("1040",),
        support_sheets=(
            "Dependents",
            "Earned Inc WS",
            "Lines 5ab (Simplified Meth WS)",
            "Line 6ab (SS Bene)",
            "Line 12 (SD Depnts)",
            "Line 12e (SD Depnts)",
            "Line 16 (QDCG Tax)",
            "Line 27 (EIC)",
            "Line 27a (EIC)",
        ),
        local_files=("F1040.ts", "F1040SR.ts"),
        local_test_files=("f1040.test.ts",),
        notes="Main individual return plus the supporting workpapers that feed core line items."
    ),
    FormGroupSpec(
        canonical_form="Schedule 1",
        primary_sheets=("Sch. 1",),
        support_sheets=("Sch 1, Line 1 (SALT)", "Sch. 1, Line 20 (IRA)", "Sch. 1, Line 21 (Stu Loan)"),
        local_files=("Schedule1.ts",),
        local_test_files=(),
        notes="Additional income and adjustments to income plus the workbook line-item worksheets."
    ),
    FormGroupSpec(
        canonical_form="Schedule 1-A",
        primary_sheets=("Sch. 1-A",),
        support_sheets=(),
        local_files=("Schedule1A.ts",),
        local_test_files=("obbba2025Law.test.ts",),
        notes="2025 deduction expansion workbook and local OBBBA implementation."
    ),
    FormGroupSpec(
        canonical_form="Schedule A",
        primary_sheets=("Sch. A",),
        support_sheets=(),
        local_files=("ScheduleA.ts",),
        local_test_files=("ScheduleA.test.ts",),
        notes="Itemized deductions."
    ),
    FormGroupSpec(
        canonical_form="Schedule B",
        primary_sheets=("Sch. B",),
        support_sheets=(),
        local_files=("ScheduleB.ts",),
        local_test_files=(),
        notes="Interest and ordinary dividends."
    ),
    FormGroupSpec(
        canonical_form="Schedule C",
        primary_sheets=("Sch. C",),
        support_sheets=(),
        local_files=("ScheduleC.ts",),
        local_test_files=(),
        notes="Business profit or loss."
    ),
    FormGroupSpec(
        canonical_form="Schedule D",
        primary_sheets=("Sch. D",),
        support_sheets=("Sch. D WS",),
        local_files=("ScheduleD.ts",),
        local_test_files=("ScheduleD.test.ts",),
        notes="Capital gains and losses, including the workbook side calculations."
    ),
    FormGroupSpec(
        canonical_form="Schedule E",
        primary_sheets=("Sch. E",),
        support_sheets=("Sch. E (2)",),
        local_files=("ScheduleE.ts",),
        local_test_files=("ScheduleEParity.test.ts",),
        notes="Supplemental income and loss, including the second workbook sheet."
    ),
    FormGroupSpec(
        canonical_form="Schedule F",
        primary_sheets=("Sch. F",),
        support_sheets=(),
        local_files=("ScheduleF.ts",),
        local_test_files=(),
        notes="Farming income and loss."
    ),
    FormGroupSpec(
        canonical_form="Schedule SE",
        primary_sheets=("Sch. SE",),
        support_sheets=(),
        local_files=("ScheduleSE.ts",),
        local_test_files=(),
        notes="Self-employment tax."
    ),
    FormGroupSpec(
        canonical_form="6251",
        primary_sheets=("6251",),
        support_sheets=(),
        local_files=("F6251.ts",),
        local_test_files=("f6251.test.ts",),
        notes="Alternative minimum tax."
    ),
    FormGroupSpec(
        canonical_form="8812",
        primary_sheets=("8812",),
        support_sheets=(),
        local_files=("Schedule8812.ts",),
        local_test_files=("Schedule8812.test.ts",),
        notes="Child tax credit and related worksheet coverage."
    ),
    FormGroupSpec(
        canonical_form="8949",
        primary_sheets=("8949A", "8949B", "8949C", "8949D"),
        support_sheets=(),
        local_files=("F8949.ts",),
        local_test_files=(),
        notes="Sales and dispositions of capital assets."
    ),
    FormGroupSpec(
        canonical_form="8995",
        primary_sheets=("8995",),
        support_sheets=(),
        local_files=("F8995.ts", "F8995A.ts"),
        local_test_files=("obbba2025Law.test.ts",),
        notes="QBI deduction workbook and local 8995/8995-A implementation."
    ),
    FormGroupSpec(
        canonical_form="EIC",
        primary_sheets=("Line 27 (EIC)", "Line 27a (EIC)"),
        support_sheets=("EIC Table",),
        local_files=("ScheduleEIC.ts",),
        local_test_files=("ScheduleEIC.test.ts",),
        notes="Earned income credit worksheets and supporting table."
    ),
)


def build_lookup() -> dict[str, FormGroupSpec]:
    lookup: dict[str, FormGroupSpec] = {}
    for spec in FORM_GROUPS:
        for sheet in (*spec.primary_sheets, *spec.support_sheets):
            lookup[sheet] = spec
    return lookup


SHEET_LOOKUP = build_lookup()


def normalize_sheet_name(sheet_name: str) -> str:
    return " ".join(sheet_name.split())


def classify_sheet(sheet_name: str) -> tuple[str | None, str]:
    normalized = normalize_sheet_name(sheet_name)
    if normalized == "1040-SR":
        return None, "auxiliary"
    if normalized in SHEET_LOOKUP:
        spec = SHEET_LOOKUP[normalized]
        if normalized in spec.primary_sheets:
            return spec.canonical_form, "target"
        return spec.canonical_form, "support"
    if normalized.startswith("1099-"):
        return None, "excluded"
    if normalized.startswith("Sch. 1") or normalized.startswith("Sch 1,"):
        return "Schedule 1", "support"
    if normalized.startswith("8949"):
        return "8949", "target"
    if "EIC" in normalized:
        return "EIC", "support"
    return None, "excluded"


def count_sheet_metrics(ws) -> tuple[int, int]:
    cell_count = 0
    formula_count = 0
    for row in ws.iter_rows():
        for cell in row:
            value = cell.value
            if value is None:
                continue
            cell_count += 1
            if getattr(cell, "data_type", None) == "f":
                formula_count += 1
            elif isinstance(value, str) and value.startswith("="):
                formula_count += 1
    return cell_count, formula_count


def workbook_label_from_path(path: Path) -> str:
    if path.name.startswith("24_"):
        return "2024"
    if path.name.startswith("25_"):
        return "2025"
    if path.name.startswith("26_"):
        return "2026"
    return path.stem


def analyze_workbook(path: Path) -> WorkbookSummary:
    workbook = load_workbook(path, read_only=True, data_only=False)
    sheets: list[SheetInventoryRow] = []
    total_cells = 0
    total_formulas = 0
    target_sheet_count = 0
    support_sheet_count = 0
    excluded_sheet_count = 0

    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        cell_count, formula_count = count_sheet_metrics(ws)
        canonical_form, role = classify_sheet(sheet_name)
        if role == "target":
            target_sheet_count += 1
        elif role == "support":
            support_sheet_count += 1
        else:
            excluded_sheet_count += 1
        total_cells += cell_count
        total_formulas += formula_count
        sheets.append(
            SheetInventoryRow(
                workbook_label=workbook_label_from_path(path),
                workbook_path=str(path),
                sheet_name=sheet_name,
                canonical_form=canonical_form,
                role=role,
                cell_count=cell_count,
                formula_count=formula_count,
                formula_density=round((formula_count / cell_count) if cell_count else 0.0, 4),
            )
        )

    return WorkbookSummary(
        workbook_label=workbook_label_from_path(path),
        workbook_path=str(path),
        sheet_count=len(workbook.sheetnames),
        target_sheet_count=target_sheet_count,
        support_sheet_count=support_sheet_count,
        excluded_sheet_count=excluded_sheet_count,
        total_cells=total_cells,
        total_formulas=total_formulas,
        sheets=sheets,
    )


def file_exists(name: str) -> bool:
    return (ROOT / "src/forms/Y2025/irsForms" / name).exists()


def make_form_row(spec: FormGroupSpec, summaries: dict[str, WorkbookSummary]) -> FormCoverageRow:
    local_present = all(file_exists(name) for name in spec.local_files)
    workbook_values = []
    years_present: list[str] = []
    years_missing: list[str] = []

    for year_label in ("2024", "2025", "2026"):
        summary = summaries[year_label]
        present_sheets: list[str] = []
        cell_count = 0
        formula_count = 0
        for sheet in summary.sheets:
            if sheet.canonical_form == spec.canonical_form:
                # Only count the sheets that belong to the target group.
                if sheet.sheet_name in spec.primary_sheets or sheet.sheet_name in spec.support_sheets:
                    present_sheets.append(sheet.sheet_name)
                    cell_count += sheet.cell_count
                    formula_count += sheet.formula_count
        if present_sheets:
            years_present.append(year_label)
        else:
            years_missing.append(year_label)
        workbook_values.append((present_sheets, cell_count, formula_count))

    return FormCoverageRow(
        canonical_form=spec.canonical_form,
        local_files=list(spec.local_files),
        local_test_files=list(spec.local_test_files),
        local_file_count=len(spec.local_files),
        local_present=local_present,
        workbook_2024_present=bool(workbook_values[0][0]),
        workbook_2024_sheet_names=workbook_values[0][0],
        workbook_2024_cell_count=workbook_values[0][1],
        workbook_2024_formula_count=workbook_values[0][2],
        workbook_2025_present=bool(workbook_values[1][0]),
        workbook_2025_sheet_names=workbook_values[1][0],
        workbook_2025_cell_count=workbook_values[1][1],
        workbook_2025_formula_count=workbook_values[1][2],
        workbook_2026_present=bool(workbook_values[2][0]),
        workbook_2026_sheet_names=workbook_values[2][0],
        workbook_2026_cell_count=workbook_values[2][1],
        workbook_2026_formula_count=workbook_values[2][2],
        years_present=years_present,
        years_missing=years_missing,
        notes=spec.notes,
    )


def to_serializable_form_row(row: FormCoverageRow) -> dict[str, object]:
    payload = asdict(row)
    payload["workbook_2024_formula_density"] = round(
        row.workbook_2024_formula_count / row.workbook_2024_cell_count, 4
    ) if row.workbook_2024_cell_count else 0.0
    payload["workbook_2025_formula_density"] = round(
        row.workbook_2025_formula_count / row.workbook_2025_cell_count, 4
    ) if row.workbook_2025_cell_count else 0.0
    payload["workbook_2026_formula_density"] = round(
        row.workbook_2026_formula_count / row.workbook_2026_cell_count, 4
    ) if row.workbook_2026_cell_count else 0.0
    payload["year_count_present"] = len(row.years_present)
    payload["year_count_missing"] = len(row.years_missing)
    return payload


def write_json(path: Path, payload: object) -> None:
    path.write_text(json.dumps(payload, indent=2, sort_keys=False) + "\n", encoding="utf-8")


def write_csv(path: Path, rows: Sequence[dict[str, object]]) -> None:
    if not rows:
        path.write_text("", encoding="utf-8")
        return
    fieldnames = list(rows[0].keys())
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            serial = {}
            for key, value in row.items():
                if isinstance(value, list):
                    serial[key] = " | ".join(str(item) for item in value)
                else:
                    serial[key] = value
            writer.writerow(serial)


def build_inventory_payload(summaries: Sequence[WorkbookSummary]) -> dict[str, object]:
    return {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "root": str(ROOT),
        "workbooks": [
            {
                "workbook_label": summary.workbook_label,
                "workbook_path": summary.workbook_path,
                "sheet_count": summary.sheet_count,
                "target_sheet_count": summary.target_sheet_count,
                "support_sheet_count": summary.support_sheet_count,
                "excluded_sheet_count": summary.excluded_sheet_count,
                "total_cells": summary.total_cells,
                "total_formulas": summary.total_formulas,
                "sheets": [asdict(sheet) for sheet in summary.sheets],
            }
            for summary in summaries
        ],
    }


def build_matrix_payload(summaries: dict[str, WorkbookSummary]) -> list[dict[str, object]]:
    rows = [to_serializable_form_row(make_form_row(spec, summaries)) for spec in FORM_GROUPS]
    return rows


def main(argv: Sequence[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description="Build 1040 workbook parity inventory and matrix outputs."
    )
    parser.add_argument(
        "--workbook",
        action="append",
        dest="workbooks",
        help="Path to a workbook. May be repeated. Defaults to the three local 1040 workbooks.",
    )
    parser.add_argument(
        "--output-dir",
        default=str(DEFAULT_OUTPUT_DIR),
        help="Directory to write the JSON/CSV artifacts into.",
    )
    args = parser.parse_args(argv)

    warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

    workbook_paths = [Path(p) for p in (args.workbooks or [str(path) for path in DEFAULT_WORKBOOKS])]
    for path in workbook_paths:
        if not path.exists():
            raise FileNotFoundError(f"Workbook not found: {path}")

    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    summaries = [analyze_workbook(path) for path in workbook_paths]
    summary_by_label = {summary.workbook_label: summary for summary in summaries}
    matrix_rows = build_matrix_payload(summary_by_label)
    inventory_payload = build_inventory_payload(summaries)

    inventory_json = output_dir / "1040_workbook_parity_inventory.json"
    matrix_json = output_dir / "1040_workbook_parity_matrix.json"
    matrix_csv = output_dir / "1040_workbook_parity_matrix.csv"

    write_json(inventory_json, inventory_payload)
    write_json(matrix_json, matrix_rows)
    write_csv(matrix_csv, matrix_rows)

    print(f"Wrote inventory: {inventory_json}")
    print(f"Wrote matrix JSON: {matrix_json}")
    print(f"Wrote matrix CSV: {matrix_csv}")
    print("")
    for summary in summaries:
        print(
            f"{summary.workbook_label}: sheets={summary.sheet_count} "
            f"targets={summary.target_sheet_count} support={summary.support_sheet_count} "
            f"excluded={summary.excluded_sheet_count} formulas={summary.total_formulas}"
        )
    print("")
    for row in matrix_rows:
        print(
            f"{row['canonical_form']}: local={row['local_present']} "
            f"years_present={row['years_present']} "
            f"2024_forms={row['workbook_2024_sheet_names']} "
            f"2025_forms={row['workbook_2025_sheet_names']} "
            f"2026_forms={row['workbook_2026_sheet_names']}"
        )

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
