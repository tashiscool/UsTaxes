#!/usr/bin/env python3
"""Deep audit for 1040-family workbooks against local implementations.

Outputs:
- sheet inventory / dependency summary
- named range inventory
- formula-by-formula trace
- coverage mapping to UsTaxes, direct-file-easy-webui, and backend-cloudflare
"""

from __future__ import annotations

import csv
import json
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from openpyxl.formula.tokenizer import Tokenizer


ROOT = Path("/Users/tkhan/IdeaProjects/taxes")
USTAXES_ROOT = ROOT / "UsTaxes"
DIRECT_FILE_ROOT = ROOT / "direct-file-easy-webui"
BACKEND_ROOT = USTAXES_ROOT / "backend-cloudflare"
DOWNLOADS = Path("/Users/tkhan/Downloads")
DEFAULT_OUTPUT_DIR = USTAXES_ROOT / "docs" / "1040_workbook_deep_audit"


WORKBOOKS = {
    "2024": DOWNLOADS / "24_1040.xlsx",
    "2025": DOWNLOADS / "25_1040.xlsx",
    "2026_forecast": DOWNLOADS / "26_1040 Forecast (Rev. 41).xlsx",
}


@dataclass(frozen=True)
class SheetMapping:
    logical_area: str
    ustaxes_files: tuple[str, ...]
    direct_file_files: tuple[str, ...]
    backend_files: tuple[str, ...]
    notes: str


COMMON_BACKEND_FILES = (
    "backend-cloudflare/src/services/taxCalculationService.ts",
    "backend-cloudflare/src/services/appSessionService.ts",
)


SHEET_MAPPINGS: dict[str, SheetMapping] = {
    "Coversheet": SheetMapping("metadata", (), (), (), "Workbook metadata only; no direct tax-engine parity target."),
    "Table of Contents": SheetMapping("metadata", (), (), (), "Navigation sheet only."),
    "Instructions": SheetMapping("instructions", (), (), (), "Instructional text, not computation."),
    "W-2s": SheetMapping("w2_intake", ("src/forms/Y2025/irsForms/F1040.ts",), (), COMMON_BACKEND_FILES, "Wage withholding intake that feeds Form 1040, 8959, dependent care, and credits."),
    "1099-INT": SheetMapping("1099_interest", ("src/forms/Y2025/irsForms/ScheduleB.ts", "src/forms/Y2025/irsForms/F1040.ts"), ("direct-file/backend/src/main/resources/tax/scheduleB.xml",), COMMON_BACKEND_FILES, "Interest income intake."),
    "1099-DIV": SheetMapping("1099_dividend", ("src/forms/Y2025/irsForms/ScheduleB.ts", "src/forms/Y2025/irsForms/ScheduleD.ts", "src/forms/Y2025/irsForms/F1040.ts"), ("direct-file/backend/src/main/resources/tax/scheduleB.xml", "direct-file/backend/src/main/resources/tax/scheduleD.xml"), COMMON_BACKEND_FILES, "Dividend intake including qualified dividends/capital gain distributions."),
    "1099-G": SheetMapping("1099_g", ("src/forms/Y2025/irsForms/Schedule1.ts", "src/forms/Y2025/irsForms/F1040.ts"), (), COMMON_BACKEND_FILES, "Unemployment and state refund intake."),
    "1099-MISC": SheetMapping("1099_misc", ("src/forms/Y2025/irsForms/ScheduleC.ts", "src/forms/Y2025/irsForms/ScheduleE.ts"), (), COMMON_BACKEND_FILES, "Miscellaneous income / royalties intake."),
    "1099-NEC": SheetMapping("1099_nec", ("src/forms/Y2025/irsForms/ScheduleC.ts", "src/forms/Y2025/irsForms/ScheduleSE.ts"), (), COMMON_BACKEND_FILES, "Self-employment intake."),
    "1099-R": SheetMapping("1099_r", ("src/forms/Y2025/irsForms/F1040.ts", "src/forms/Y2025/irsForms/F8606.ts"), ("direct-file/backend/src/main/resources/tax/form8606.xml",), COMMON_BACKEND_FILES, "Retirement distributions and basis interactions."),
    "Lines 4ab (IRA Dist.)": SheetMapping("ira_distribution_worksheet", ("src/forms/Y2025/irsForms/F1040.ts", "src/forms/Y2025/irsForms/F8606.ts"), ("direct-file/backend/src/main/resources/tax/form8606.xml",), COMMON_BACKEND_FILES, "IRA taxable amount worksheet."),
    "SSA-1099": SheetMapping("social_security_intake", ("src/forms/Y2025/irsForms/F1040.ts",), (), COMMON_BACKEND_FILES, "Social Security benefits intake."),
    "1040": SheetMapping("form_1040", ("src/forms/Y2025/irsForms/F1040.ts",), ("direct-file/backend/src/main/resources/tax/form1040SR.xml",), COMMON_BACKEND_FILES, "Primary return and totals."),
    "1040-SR": SheetMapping("form_1040_sr", ("src/forms/Y2025/irsForms/F1040SR.ts", "src/forms/Y2025/irsForms/F1040.ts"), ("direct-file/backend/src/main/resources/tax/form1040SR.xml",), COMMON_BACKEND_FILES, "Senior layout / parity with main 1040 engine."),
    "Dependents": SheetMapping("dependents", ("src/forms/Y2025/irsForms/ScheduleEIC.ts", "src/forms/Y2025/irsForms/Schedule8812.ts", "src/forms/Y2025/irsForms/F1040.ts"), ("direct-file/backend/src/main/resources/tax/scheduleEIC.xml", "direct-file/backend/src/main/resources/tax/form8812.xml", "direct-file/backend/src/main/resources/tax/schedule8812.xml"), COMMON_BACKEND_FILES, "Dependent roster driving EIC/CTC."),
    "Sch. 1": SheetMapping("schedule_1", ("src/forms/Y2025/irsForms/Schedule1.ts",), (), COMMON_BACKEND_FILES, "Additional income and adjustments."),
    "Sch. 1-A": SheetMapping("schedule_1a", ("src/forms/Y2025/irsForms/Schedule1A.ts",), ("direct-file/backend/src/main/resources/tax/schedule1A.xml",), COMMON_BACKEND_FILES, "OBBBA deduction schedule for TY2025+."),
    "Sch 1, Line 1 (SALT)": SheetMapping("salt_worksheet", ("src/forms/Y2025/irsForms/ScheduleA.ts", "src/forms/Y2025/irsForms/Schedule1A.ts"), ("direct-file/backend/src/main/resources/tax/scheduleA.xml", "direct-file/backend/src/main/resources/tax/schedule1A.xml"), COMMON_BACKEND_FILES, "SALT itemized / special deduction worksheet."),
    "Sch. 1, Line 20 (IRA)": SheetMapping("ira_adjustment_worksheet", ("src/forms/Y2025/irsForms/Schedule1.ts", "src/forms/Y2025/irsForms/F8606.ts", "src/forms/Y2025/irsForms/F8880.ts"), ("direct-file/backend/src/main/resources/tax/form8606.xml", "direct-file/backend/src/main/resources/tax/form8880.xml"), COMMON_BACKEND_FILES, "IRA deduction / saver’s credit interactions."),
    "Sch. 1, Line 21 (Stu Loan)": SheetMapping("student_loan_interest_worksheet", ("src/forms/Y2025/irsForms/Schedule1.ts",), (), COMMON_BACKEND_FILES, "Student loan interest deduction."),
    "Sch. 2": SheetMapping("schedule_2", ("src/forms/Y2025/irsForms/F6251.ts", "src/forms/Y2025/irsForms/F8959.ts", "src/forms/Y2025/irsForms/F8960.ts", "src/forms/Y2025/irsForms/F8801.ts"), ("direct-file/backend/src/main/resources/tax/form6251.xml", "direct-file/backend/src/main/resources/tax/form8959.xml", "direct-file/backend/src/main/resources/tax/form8960.xml", "direct-file/backend/src/main/resources/tax/form8801.xml"), COMMON_BACKEND_FILES, "Additional taxes."),
    "Sch. 3": SheetMapping("schedule_3", ("src/forms/Y2025/irsForms/F2441.ts", "src/forms/Y2025/irsForms/F8863.ts", "src/forms/Y2025/irsForms/F5695.ts", "src/forms/Y2025/irsForms/F8801.ts"), ("direct-file/backend/src/main/resources/tax/form2441.xml", "direct-file/backend/src/main/resources/tax/form8863.xml", "direct-file/backend/src/main/resources/tax/form5695.xml", "direct-file/backend/src/main/resources/tax/form8801.xml"), COMMON_BACKEND_FILES, "Nonrefundable and refundable credits."),
    "Earned Inc WS": SheetMapping("earned_income_worksheet", ("src/forms/Y2025/irsForms/ScheduleEIC.ts", "src/forms/Y2025/irsForms/F1040.ts"), ("direct-file/backend/src/main/resources/tax/scheduleEIC.xml",), COMMON_BACKEND_FILES, "Earned income worksheet."),
    "Lines 5ab (Simplified Meth WS)": SheetMapping("pension_simplified_method", ("src/forms/Y2025/irsForms/F1040.ts",), (), COMMON_BACKEND_FILES, "Pension/annuity simplified method support."),
    "Line 6ab (SS Bene)": SheetMapping("social_security_worksheet", ("src/forms/Y2025/irsForms/F1040.ts",), (), COMMON_BACKEND_FILES, "Taxable Social Security benefits worksheet."),
    "Line 12 (SD Depnts)": SheetMapping("standard_deduction_dependents", ("src/forms/Y2024/irsForms/F1040.ts",), (), COMMON_BACKEND_FILES, "2024 dependent standard deduction worksheet."),
    "Line 12e (SD Depnts)": SheetMapping("standard_deduction_dependents", ("src/forms/Y2025/irsForms/F1040.ts",), (), COMMON_BACKEND_FILES, "2025+ dependent standard deduction worksheet."),
    "Line 16 (QDCG Tax)": SheetMapping("qualified_dividend_cap_gains_tax", ("src/forms/Y2025/irsForms/F1040.ts", "src/forms/Y2025/irsForms/ScheduleD.ts"), ("direct-file/backend/src/main/resources/tax/scheduleD.xml",), COMMON_BACKEND_FILES, "Qualified dividends/capital gains tax worksheet."),
    "Line 27 (EIC)": SheetMapping("eitc", ("src/forms/Y2024/irsForms/ScheduleEIC.ts", "src/forms/Y2024/irsForms/F1040.ts"), ("direct-file/backend/src/main/resources/tax/scheduleEIC.xml",), COMMON_BACKEND_FILES, "2024 EIC worksheet."),
    "Line 27a (EIC)": SheetMapping("eitc", ("src/forms/Y2025/irsForms/ScheduleEIC.ts", "src/forms/Y2025/irsForms/F1040.ts"), ("direct-file/backend/src/main/resources/tax/scheduleEIC.xml",), COMMON_BACKEND_FILES, "2025+ EIC worksheet."),
    "Sch. A": SheetMapping("schedule_a", ("src/forms/Y2025/irsForms/ScheduleA.ts",), ("direct-file/backend/src/main/resources/tax/scheduleA.xml",), COMMON_BACKEND_FILES, "Itemized deductions."),
    "Sch. B": SheetMapping("schedule_b", ("src/forms/Y2025/irsForms/ScheduleB.ts",), ("direct-file/backend/src/main/resources/tax/scheduleB.xml",), COMMON_BACKEND_FILES, "Interest and dividends."),
    "Sch. C": SheetMapping("schedule_c", ("src/forms/Y2025/irsForms/ScheduleC.ts", "src/forms/Y2025/irsForms/ScheduleSE.ts", "src/forms/Y2025/irsForms/F8995.ts", "src/forms/Y2025/irsForms/F8995A.ts"), ("direct-file/backend/src/main/resources/tax/scheduleC.xml", "direct-file/backend/src/main/resources/tax/scheduleSE.xml", "direct-file/backend/src/main/resources/tax/form8995.xml", "direct-file/backend/src/main/resources/tax/form8995A.xml"), COMMON_BACKEND_FILES, "Business income and QBI source."),
    "Sch. D": SheetMapping("schedule_d", ("src/forms/Y2025/irsForms/ScheduleD.ts", "src/forms/Y2025/irsForms/F8949.ts"), ("direct-file/backend/src/main/resources/tax/scheduleD.xml", "direct-file/backend/src/main/resources/tax/form8949.xml"), COMMON_BACKEND_FILES, "Capital gains summary."),
    "Sch. D WS": SheetMapping("schedule_d_worksheet", ("src/forms/Y2025/irsForms/ScheduleD.ts", "src/forms/Y2025/irsForms/F8949.ts", "src/forms/Y2025/irsForms/F1040.ts"), ("direct-file/backend/src/main/resources/tax/scheduleD.xml", "direct-file/backend/src/main/resources/tax/form8949.xml"), COMMON_BACKEND_FILES, "Capital gain tax and carryover worksheet."),
    "Sch. E": SheetMapping("schedule_e", ("src/forms/Y2025/irsForms/ScheduleE.ts", "src/forms/Y2025/irsForms/F8995.ts", "src/forms/Y2025/irsForms/F8995A.ts"), ("direct-file/backend/src/main/resources/tax/scheduleE.xml", "direct-file/backend/src/main/resources/tax/form8995.xml", "direct-file/backend/src/main/resources/tax/form8995A.xml"), COMMON_BACKEND_FILES, "Rental, royalty, K-1 pass-through income."),
    "Sch. E (2)": SheetMapping("schedule_e_page_2", ("src/forms/Y2025/irsForms/ScheduleE.ts", "src/forms/Y2025/irsForms/F8995.ts", "src/forms/Y2025/irsForms/F8995A.ts"), ("direct-file/backend/src/main/resources/tax/scheduleE.xml", "direct-file/backend/src/main/resources/tax/form8995.xml", "direct-file/backend/src/main/resources/tax/form8995A.xml"), COMMON_BACKEND_FILES, "Second Schedule E page / overflow."),
    "Sch. F": SheetMapping("schedule_f", ("src/forms/Y2025/irsForms/ScheduleF.ts", "src/forms/Y2025/irsForms/ScheduleSE.ts"), ("direct-file/backend/src/main/resources/tax/scheduleF.xml",), COMMON_BACKEND_FILES, "Farm income."),
    "Sch. R": SheetMapping("schedule_r", ("src/forms/Y2025/irsForms/ScheduleR.ts",), ("direct-file/backend/src/main/resources/tax/scheduleR.xml",), COMMON_BACKEND_FILES, "Credit for the elderly or disabled."),
    "Sch. SE": SheetMapping("schedule_se", ("src/forms/Y2025/irsForms/ScheduleSE.ts",), ("direct-file/backend/src/main/resources/tax/scheduleSE.xml",), COMMON_BACKEND_FILES, "Self-employment tax."),
    "1116": SheetMapping("form_1116", ("src/forms/Y2025/irsForms/F1116.ts",), ("direct-file/backend/src/main/resources/tax/form1116.xml",), COMMON_BACKEND_FILES, "Foreign tax credit."),
    "2210": SheetMapping("form_2210", ("src/forms/Y2025/irsForms/F2210.ts",), ("direct-file/backend/src/main/resources/tax/form2210.xml",), COMMON_BACKEND_FILES, "Underpayment of estimated tax."),
    "2441": SheetMapping("form_2441", ("src/forms/Y2025/irsForms/F2441.ts",), ("direct-file/backend/src/main/resources/tax/form2441.xml",), COMMON_BACKEND_FILES, "Dependent care credit."),
    "2555": SheetMapping("form_2555", ("src/forms/Y2025/irsForms/F2555.ts",), ("direct-file/backend/src/main/resources/tax/form2555.xml",), COMMON_BACKEND_FILES, "Foreign earned income exclusion."),
    "4137": SheetMapping("form_4137", ("src/forms/Y2025/irsForms/F4137.ts",), ("direct-file/backend/src/main/resources/tax/form4137.xml",), COMMON_BACKEND_FILES, "Social Security/Medicare tax on unreported tip income."),
    "4562": SheetMapping("form_4562", ("src/forms/Y2025/irsForms/F4562.ts",), ("direct-file/backend/src/main/resources/tax/form4562.xml",), COMMON_BACKEND_FILES, "Depreciation and section 179."),
    "5695": SheetMapping("form_5695", ("src/forms/Y2025/irsForms/F5695.ts",), ("direct-file/backend/src/main/resources/tax/form5695.xml",), COMMON_BACKEND_FILES, "Residential energy credits."),
    "6251": SheetMapping("form_6251", ("src/forms/Y2025/irsForms/F6251.ts",), ("direct-file/backend/src/main/resources/tax/form6251.xml",), COMMON_BACKEND_FILES, "Alternative minimum tax."),
    "8283": SheetMapping("form_8283", ("src/forms/Y2025/irsForms/F8283.ts",), ("direct-file/backend/src/main/resources/tax/form8283.xml",), COMMON_BACKEND_FILES, "Noncash charitable contributions."),
    "8582": SheetMapping("form_8582", ("src/forms/Y2025/irsForms/ScheduleE.ts",), (), COMMON_BACKEND_FILES, "Passive activity loss limitations that feed Schedule E."),
    "8606": SheetMapping("form_8606", ("src/forms/Y2025/irsForms/F8606.ts",), ("direct-file/backend/src/main/resources/tax/form8606.xml",), COMMON_BACKEND_FILES, "Nondeductible IRAs."),
    "8801": SheetMapping("form_8801", ("src/forms/Y2025/irsForms/F8801.ts",), ("direct-file/backend/src/main/resources/tax/form8801.xml",), COMMON_BACKEND_FILES, "Credit for prior year minimum tax."),
    "8812": SheetMapping("form_8812", ("src/forms/Y2025/irsForms/Schedule8812.ts",), ("direct-file/backend/src/main/resources/tax/form8812.xml", "direct-file/backend/src/main/resources/tax/schedule8812.xml"), COMMON_BACKEND_FILES, "Child tax credit and additional child tax credit."),
    "8829": SheetMapping("form_8829", ("src/forms/Y2025/irsForms/F8829.ts", "src/forms/Y2025/irsForms/ScheduleC.ts"), ("direct-file/backend/src/main/resources/tax/form8829.xml",), COMMON_BACKEND_FILES, "Home office expenses."),
    "8863": SheetMapping("form_8863", ("src/forms/Y2025/irsForms/F8863.ts",), ("direct-file/backend/src/main/resources/tax/form8863.xml",), COMMON_BACKEND_FILES, "Education credits."),
    "8880": SheetMapping("form_8880", ("src/forms/Y2025/irsForms/F8880.ts",), ("direct-file/backend/src/main/resources/tax/form8880.xml",), COMMON_BACKEND_FILES, "Retirement savings contributions credit."),
    "8889": SheetMapping("form_8889", ("src/forms/Y2025/irsForms/F8889.ts",), ("direct-file/backend/src/main/resources/tax/form8889.xml",), COMMON_BACKEND_FILES, "Health savings accounts."),
    "8919": SheetMapping("form_8919", ("src/forms/Y2025/irsForms/F8919.ts",), ("direct-file/backend/src/main/resources/tax/form8919.xml",), COMMON_BACKEND_FILES, "Uncollected Social Security and Medicare tax on wages."),
    "8949A": SheetMapping("form_8949", ("src/forms/Y2025/irsForms/F8949.ts", "src/forms/Y2025/irsForms/ScheduleD.ts"), ("direct-file/backend/src/main/resources/tax/form8949.xml", "direct-file/backend/src/main/resources/tax/scheduleD.xml"), COMMON_BACKEND_FILES, "8949 part A."),
    "8949B": SheetMapping("form_8949", ("src/forms/Y2025/irsForms/F8949.ts", "src/forms/Y2025/irsForms/ScheduleD.ts"), ("direct-file/backend/src/main/resources/tax/form8949.xml", "direct-file/backend/src/main/resources/tax/scheduleD.xml"), COMMON_BACKEND_FILES, "8949 part B."),
    "8949C": SheetMapping("form_8949", ("src/forms/Y2025/irsForms/F8949.ts", "src/forms/Y2025/irsForms/ScheduleD.ts"), ("direct-file/backend/src/main/resources/tax/form8949.xml", "direct-file/backend/src/main/resources/tax/scheduleD.xml"), COMMON_BACKEND_FILES, "8949 part C."),
    "8949D": SheetMapping("form_8949", ("src/forms/Y2025/irsForms/F8949.ts", "src/forms/Y2025/irsForms/ScheduleD.ts"), ("direct-file/backend/src/main/resources/tax/form8949.xml", "direct-file/backend/src/main/resources/tax/scheduleD.xml"), COMMON_BACKEND_FILES, "8949 part D."),
    "8959": SheetMapping("form_8959", ("src/forms/Y2025/irsForms/F8959.ts",), ("direct-file/backend/src/main/resources/tax/form8959.xml",), COMMON_BACKEND_FILES, "Additional Medicare tax."),
    "8960": SheetMapping("form_8960", ("src/forms/Y2025/irsForms/F8960.ts",), ("direct-file/backend/src/main/resources/tax/form8960.xml",), COMMON_BACKEND_FILES, "Net investment income tax."),
    "8962": SheetMapping("form_8962", ("src/forms/Y2025/irsForms/F8962.ts",), ("direct-file/backend/src/main/resources/tax/form8962.xml",), COMMON_BACKEND_FILES, "Premium tax credit."),
    "8995": SheetMapping("form_8995", ("src/forms/Y2025/irsForms/F8995.ts", "src/forms/Y2025/irsForms/F8995A.ts"), ("direct-file/backend/src/main/resources/tax/form8995.xml", "direct-file/backend/src/main/resources/tax/form8995A.xml"), COMMON_BACKEND_FILES, "Qualified business income deduction."),
    "EIC Table": SheetMapping("eitc_table", ("src/forms/Y2025/irsForms/ScheduleEIC.ts", "src/forms/Y2025/irsForms/F1040.ts"), ("direct-file/backend/src/main/resources/tax/scheduleEIC.xml",), COMMON_BACKEND_FILES, "Earned income credit lookup table."),
    "Tax Table": SheetMapping("tax_table", ("src/forms/Y2025/irsForms/F1040.ts",), ("direct-file/backend/src/main/resources/tax/form1040SR.xml",), COMMON_BACKEND_FILES, "Rate/table lookups for tax calculation."),
    "Changes": SheetMapping("change_log", (), (), (), "Workbook release notes / changelog."),
}


def ensure_output_dir(path: Path) -> None:
    path.mkdir(parents=True, exist_ok=True)


def norm_rel(path: Path, root: Path) -> str:
    return str(path.relative_to(root))


def file_exists(root: Path, rel_paths: Iterable[str]) -> list[str]:
    return [path for path in rel_paths if (root / path).exists()]


def tokenize_references(formula: str) -> list[str]:
    references: list[str] = []
    try:
        tokenizer = Tokenizer(formula)
    except Exception:
        return references

    for token in tokenizer.items:
        if token.subtype == "RANGE" and token.value:
            references.append(token.value)
    return references


def reference_sheets(refs: Iterable[str]) -> list[str]:
    sheets: list[str] = []
    seen: set[str] = set()
    for ref in refs:
        if "!" not in ref:
            continue
        sheet = ref.split("!", 1)[0].strip("'")
        if sheet not in seen:
            seen.add(sheet)
            sheets.append(sheet)
    return sheets


def coverage_status(mapping: SheetMapping) -> str:
    if not mapping.ustaxes_files and not mapping.direct_file_files and not mapping.backend_files:
        return "informational_only"
    if not mapping.ustaxes_files:
        return "missing_ustaxes_mapping"
    if not mapping.direct_file_files:
        return "missing_direct_file_mapping"
    return "mapped"


def extract_defined_names(wb) -> list[dict[str, object]]:
    entries: list[dict[str, object]] = []
    for name, defn in wb.defined_names.items():
        value = getattr(defn, "value", None) or getattr(defn, "attr_text", None)
        destinations = []
        try:
            for title, coord in defn.destinations:
                destinations.append({"sheet": title, "coord": coord})
        except Exception:
            pass
        entries.append(
            {
                "name": name,
                "value": value,
                "destinations": destinations,
            }
        )
    return entries


def audit_workbook(label: str, workbook_path: Path) -> dict[str, object]:
    wb = load_workbook(workbook_path, data_only=False, read_only=True)

    sheet_rows: list[dict[str, object]] = []
    formula_rows: list[dict[str, object]] = []
    for ws in wb.worksheets:
        mapping = SHEET_MAPPINGS.get(
            ws.title,
            SheetMapping(
                logical_area="unmapped",
                ustaxes_files=(),
                direct_file_files=(),
                backend_files=(),
                notes="No manual mapping yet.",
            ),
        )
        nonempty = 0
        formulas = 0
        per_sheet_refs: dict[str, int] = {}

        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if value is None:
                    continue
                nonempty += 1
                if isinstance(value, str) and value.startswith("="):
                    formulas += 1
                    refs = tokenize_references(value)
                    ref_sheets = reference_sheets(refs)
                    for ref_sheet in ref_sheets:
                        per_sheet_refs[ref_sheet] = per_sheet_refs.get(ref_sheet, 0) + 1
                    formula_rows.append(
                        {
                            "workbook_label": label,
                            "workbook_file": workbook_path.name,
                            "sheet": ws.title,
                            "cell": cell.coordinate,
                            "formula": value,
                            "references": json.dumps(refs),
                            "reference_sheets": json.dumps(ref_sheets),
                            "logical_area": mapping.logical_area,
                            "ustaxes_files": json.dumps(file_exists(USTAXES_ROOT, mapping.ustaxes_files)),
                            "direct_file_files": json.dumps(file_exists(DIRECT_FILE_ROOT, mapping.direct_file_files)),
                            "backend_files": json.dumps(file_exists(USTAXES_ROOT, mapping.backend_files)),
                            "coverage_status": coverage_status(mapping),
                            "notes": mapping.notes,
                        }
                    )

        sheet_rows.append(
            {
                "workbook_label": label,
                "workbook_file": workbook_path.name,
                "sheet": ws.title,
                "max_row": ws.max_row,
                "max_col": ws.max_column,
                "nonempty_cells": nonempty,
                "formula_cells": formulas,
                "logical_area": mapping.logical_area,
                "ustaxes_files": json.dumps(file_exists(USTAXES_ROOT, mapping.ustaxes_files)),
                "direct_file_files": json.dumps(file_exists(DIRECT_FILE_ROOT, mapping.direct_file_files)),
                "backend_files": json.dumps(file_exists(USTAXES_ROOT, mapping.backend_files)),
                "cross_sheet_reference_counts": json.dumps(per_sheet_refs, sort_keys=True),
                "coverage_status": coverage_status(mapping),
                "notes": mapping.notes,
            }
        )

    defined_names = extract_defined_names(wb)
    name_rows: list[dict[str, object]] = []
    for entry in defined_names:
        destinations = entry["destinations"]
        first_sheet = destinations[0]["sheet"] if destinations else None
        mapping = SHEET_MAPPINGS.get(
            first_sheet or "",
            SheetMapping(
                logical_area="named_range_unmapped",
                ustaxes_files=(),
                direct_file_files=(),
                backend_files=(),
                notes="Named range target could not be mapped to a known sheet.",
            ),
        )
        name_rows.append(
            {
                "workbook_label": label,
                "workbook_file": workbook_path.name,
                "defined_name": entry["name"],
                "value": entry["value"],
                "destinations": json.dumps(destinations),
                "logical_area": mapping.logical_area,
                "ustaxes_files": json.dumps(file_exists(USTAXES_ROOT, mapping.ustaxes_files)),
                "direct_file_files": json.dumps(file_exists(DIRECT_FILE_ROOT, mapping.direct_file_files)),
                "backend_files": json.dumps(file_exists(USTAXES_ROOT, mapping.backend_files)),
                "coverage_status": coverage_status(mapping),
                "notes": mapping.notes,
            }
        )

    return {
        "sheet_rows": sheet_rows,
        "formula_rows": formula_rows,
        "name_rows": name_rows,
    }


def write_csv(rows: list[dict[str, object]], path: Path) -> None:
    if not rows:
        path.write_text("")
        return
    with path.open("w", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)


def write_json(payload: object, path: Path) -> None:
    path.write_text(json.dumps(payload, indent=2) + "\n")


def build_summary_markdown(
    all_sheets: list[dict[str, object]],
    all_names: list[dict[str, object]],
    output_path: Path,
) -> None:
    lines = [
        "# 1040 Workbook Deep Audit",
        "",
        "This report inventories the 2024, 2025, and 2026 forecast workbooks sheet-by-sheet,",
        "tracks formula cells, extracts defined names, and maps workbook areas to local",
        "`UsTaxes`, `direct-file-easy-webui`, and `backend-cloudflare` implementations.",
        "",
        "## Quick findings",
        "",
    ]

    for workbook_label in ("2024", "2025", "2026_forecast"):
        sheets = [row for row in all_sheets if row["workbook_label"] == workbook_label]
        formulas = sum(int(row["formula_cells"]) for row in sheets)
        lines.append(
            f"- `{workbook_label}`: {len(sheets)} sheets, {formulas:,} formula cells, {sum(1 for row in all_names if row['workbook_label'] == workbook_label):,} defined names."
        )

    lines.extend(
        [
            "",
            "## Coverage interpretation",
            "",
            "- `mapped`: workbook sheet/name has an explicit mapping to local code artifacts.",
            "- `missing_direct_file_mapping`: the workbook area is mapped in `UsTaxes`, but no direct-file mapping was configured in this audit.",
            "- `missing_ustaxes_mapping`: direct-file/backend hooks exist but no `UsTaxes` mapping was configured.",
            "- `informational_only`: workbook tab is metadata, navigation, or instructions rather than tax computation.",
            "",
            "## Evidence-based caution",
            "",
            "- A workbook-to-code mapping is not proof of line-by-line parity.",
            "- Full parity requires comparing output values and logic branches for each workbook area against the implementations.",
            "- The formula trace CSV is the starting point for that proof, not the proof by itself.",
            "",
        ]
    )

    output_path.write_text("\n".join(lines) + "\n")


def main() -> int:
    ensure_output_dir(DEFAULT_OUTPUT_DIR)

    all_sheets: list[dict[str, object]] = []
    all_formulas: list[dict[str, object]] = []
    all_names: list[dict[str, object]] = []

    for label, workbook_path in WORKBOOKS.items():
        result = audit_workbook(label, workbook_path)
        all_sheets.extend(result["sheet_rows"])
        all_formulas.extend(result["formula_rows"])
        all_names.extend(result["name_rows"])

    write_csv(all_sheets, DEFAULT_OUTPUT_DIR / "sheet_inventory.csv")
    write_csv(all_names, DEFAULT_OUTPUT_DIR / "defined_names.csv")
    write_csv(all_formulas, DEFAULT_OUTPUT_DIR / "formula_trace.csv")
    write_json(all_sheets, DEFAULT_OUTPUT_DIR / "sheet_inventory.json")
    write_json(all_names, DEFAULT_OUTPUT_DIR / "defined_names.json")
    write_json(
        {
            "row_count": len(all_formulas),
            "sample": all_formulas[:500],
        },
        DEFAULT_OUTPUT_DIR / "formula_trace_sample.json",
    )
    build_summary_markdown(
        all_sheets,
        all_names,
        DEFAULT_OUTPUT_DIR / "README.md",
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
