#!/usr/bin/env python3
"""Build a machine-readable business-form parity harness.

This harness extends the 1040 workbook parity approach to business/entity forms.
When private workbooks are unavailable, it inventories the local IRS MeF/ATS
materials that can still anchor parity work and records the workbook gap
explicitly instead of pretending coverage exists.
"""

from __future__ import annotations

import argparse
import csv
import json
import warnings
from dataclasses import asdict, dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Iterable, Sequence

try:
    from openpyxl import load_workbook  # type: ignore
except Exception:  # pragma: no cover - dependency is expected but optional here
    load_workbook = None

try:
    import xlrd  # type: ignore
except Exception:  # pragma: no cover - optional for legacy .xls materials
    xlrd = None


ROOT = Path('/Users/tkhan/IdeaProjects/taxes/UsTaxes')
DOWNLOADS = Path('/Users/tkhan/Downloads')
DEFAULT_OUTPUT_DIR = ROOT / 'docs'
IRS_MEF_ROOT = DOWNLOADS / 'IRS_MeF_Materials'
FIXTURE_MANIFEST = ROOT / 'src/tests/ats/business/fixtures/business_fixture_manifest.json'


@dataclass(frozen=True)
class FormSpec:
    canonical_form: str
    workbook_patterns: tuple[str, ...]
    irs_reference_patterns: tuple[str, ...]
    local_files: tuple[str, ...]
    local_test_files: tuple[str, ...]
    notes: str


@dataclass
class WorkbookDetail:
    path: str
    file_type: str
    sheet_count: int | None
    sheet_names: list[str]


@dataclass
class CoverageRow:
    canonical_form: str
    workbook_present: bool
    workbook_paths: list[str]
    workbook_sheet_names: list[str]
    workbook_sheet_count: int
    irs_reference_present: bool
    irs_reference_count: int
    irs_reference_sample_paths: list[str]
    local_files: list[str]
    local_file_count: int
    local_present: bool
    local_test_files: list[str]
    local_test_count: int
    local_tests_present: bool
    fixture_backed: bool
    fixture_count: int
    fixture_paths: list[str]
    coverage_status: str
    notes: str
    last_audited: str


FORM_SPECS: tuple[FormSpec, ...] = (
    FormSpec(
        canonical_form='1120',
        workbook_patterns=(
            '*1120*.xlsx',
            '*1120*.xlsm',
            '*1120*.xls',
            '*corp*.xlsx',
        ),
        irs_reference_patterns=(
            'Known_Issues/*1120formfamily*.xls',
            'Test_Scenarios/TY2025/*1120*.zip',
            'Schema_Package/extracted/**/2025/IRS1120*.xsl',
            'Schema_Package/extracted/**/2025/IRS1120*.xsd',
        ),
        local_files=(
            'src/forms/Y2025/irsForms/F1120.ts',
            'src/efile/mef/businessEntitySerializer.ts',
            'backend-cloudflare/src/services/taxCalculationService.ts',
            'backend-cloudflare/src/services/mefComplianceService.ts',
        ),
        local_test_files=(
            'backend-cloudflare/test/service/businessEntityCalc.test.ts',
            'backend-cloudflare/test/service/taxCalculationService.businessEntities.test.ts',
            'src/tests/ats/business/ccorp.test.ts',
        ),
        notes='C-corporation parity should track Form 1120 computation, MeF serialization, and backend entity-return handling.',
    ),
    FormSpec(
        canonical_form='1120-S',
        workbook_patterns=(
            '*1120-S*.xlsx',
            '*1120S*.xlsx',
            '*1120-S*.xlsm',
            '*1120S*.xlsm',
            '*1120-S*.xls',
            '*1120S*.xls',
        ),
        irs_reference_patterns=(
            'Known_Issues/*1120formfamily*.xls',
            'Schema_Package/extracted/**/2025/IRS1120S*.xsl',
            'Schema_Package/extracted/**/2025/IRS1120S*.xsd',
        ),
        local_files=(
            'src/forms/Y2025/irsForms/F1120S.ts',
            'src/efile/mef/businessEntitySerializer.ts',
            'backend-cloudflare/src/services/taxCalculationService.ts',
            'backend-cloudflare/src/services/k1GenerationService.ts',
        ),
        local_test_files=(
            'backend-cloudflare/test/service/businessEntityCalc.test.ts',
            'backend-cloudflare/test/service/taxCalculationService.businessEntities.test.ts',
            'src/tests/ats/business/scorp.test.ts',
        ),
        notes='S-corp parity should include ordinary business income, Schedule K/K-1 allocation, and shareholder ownership inputs.',
    ),
    FormSpec(
        canonical_form='1065',
        workbook_patterns=(
            '*1065*.xlsx',
            '*1065*.xlsm',
            '*1065*.xls',
            '*partnership*.xlsx',
        ),
        irs_reference_patterns=(
            'Known_Issues/*1065*.xls',
            'Test_Scenarios/1065_Partnership/*.pdf',
            'Schema_Package/extracted/**/2025/IRS1065*.xsl',
            'Schema_Package/extracted/**/2025/IRS1065*.xsd',
        ),
        local_files=(
            'src/forms/Y2025/irsForms/F1065.ts',
            'src/efile/mef/businessEntitySerializer.ts',
            'backend-cloudflare/src/services/taxCalculationService.ts',
            'backend-cloudflare/src/services/k1GenerationService.ts',
        ),
        local_test_files=(
            'backend-cloudflare/test/service/businessEntityCalc.test.ts',
            'backend-cloudflare/test/service/taxCalculationService.businessEntities.test.ts',
            'src/tests/ats/business/partnership.test.ts',
        ),
        notes='Partnership parity should cover guaranteed payments, partner allocations, Schedule K/K-1, and liabilities.',
    ),
    FormSpec(
        canonical_form='1041',
        workbook_patterns=(
            '*1041*.xlsx',
            '*1041*.xlsm',
            '*1041*.xls',
            '*trust*.xlsx',
            '*estate*.xlsx',
        ),
        irs_reference_patterns=(
            'Test_Scenarios/1041_Estate_Trust/*.pdf',
            'Schema_Package/extracted/**/2025/IRS1041*.xsl',
            'Schema_Package/extracted/**/2025/IRS1041*.xsd',
        ),
        local_files=(
            'src/forms/Y2025/irsForms/F1041.ts',
            'src/forms/Y2025/irsForms/F1041A.ts',
            'src/forms/Y2025/irsForms/F1041ES.ts',
            'backend-cloudflare/src/services/taxCalculationService.ts',
            'backend-cloudflare/src/services/k1GenerationService.ts',
        ),
        local_test_files=(
            'backend-cloudflare/test/service/businessEntityCalc.test.ts',
            'backend-cloudflare/test/service/taxCalculationService.businessEntities.test.ts',
        ),
        notes='Trust and estate parity is driven by fiduciary, beneficiary, distribution, and compressed-bracket logic.',
    ),
    FormSpec(
        canonical_form='990',
        workbook_patterns=(
            '*990*.xlsx',
            '*990*.xlsm',
            '*990*.xls',
            '*nonprofit*.xlsx',
        ),
        irs_reference_patterns=(
            'Test_Scenarios/TY2025/*990*.zip',
            'Schema_Package/extracted/**/2025/IRS990*.xsl',
            'Schema_Package/extracted/**/2025/IRS990*.xsd',
        ),
        local_files=(
            'src/forms/Y2025/irsForms/F990.ts',
            'src/forms/Y2025/irsForms/F990EZ.ts',
            'src/forms/Y2025/irsForms/F990N.ts',
            'src/forms/Y2025/irsForms/F990PF.ts',
            'src/forms/Y2025/irsForms/F990T.ts',
            'backend-cloudflare/src/services/taxCalculationService.ts',
        ),
        local_test_files=(
            'backend-cloudflare/test/service/businessEntityCalc.test.ts',
            'backend-cloudflare/test/service/taxCalculationService.businessEntities.test.ts',
        ),
        notes='Nonprofit parity is partially implemented in local forms, but Cloudflare still routes Form 990-family work to expert handling.',
    ),
)


def find_matching_files(base: Path, patterns: Sequence[str]) -> list[Path]:
    matches: list[Path] = []
    seen: set[Path] = set()
    for pattern in patterns:
        for path in base.glob(pattern):
            resolved = path.resolve()
            if resolved not in seen:
                seen.add(resolved)
                matches.append(resolved)
    return sorted(matches)


def summarize_workbook(path: Path) -> WorkbookDetail:
    suffix = path.suffix.lower()
    if suffix in {'.xlsx', '.xlsm'} and load_workbook is not None:
        workbook = load_workbook(path, read_only=True, data_only=False)
        try:
            sheet_names = list(workbook.sheetnames)
            return WorkbookDetail(
                path=str(path),
                file_type=suffix,
                sheet_count=len(sheet_names),
                sheet_names=sheet_names,
            )
        finally:
            workbook.close()

    if suffix == '.xls' and xlrd is not None:
        workbook = xlrd.open_workbook(path)
        sheet_names = workbook.sheet_names()
        return WorkbookDetail(
            path=str(path),
            file_type=suffix,
            sheet_count=len(sheet_names),
            sheet_names=list(sheet_names),
        )

    warnings.warn(
        f'Could not inspect workbook structure for {path.name}; file type {suffix} requires openpyxl/xlrd support.'
    )
    return WorkbookDetail(
        path=str(path),
        file_type=suffix,
        sheet_count=None,
        sheet_names=[],
    )


def filter_local_files(paths: Iterable[str]) -> list[str]:
    return [str((ROOT / rel).resolve()) for rel in paths if (ROOT / rel).exists()]


def load_fixture_manifest() -> dict[str, list[dict[str, object]]]:
    if not FIXTURE_MANIFEST.exists():
        return {}

    payload = json.loads(FIXTURE_MANIFEST.read_text())
    fixtures_by_form: dict[str, list[dict[str, object]]] = {}
    for raw_entry in payload.get('fixtures', []):
        if not isinstance(raw_entry, dict):
            continue
        form_type = raw_entry.get('formType')
        if not isinstance(form_type, str) or not form_type:
            continue
        fixtures_by_form.setdefault(form_type, []).append(raw_entry)
    return fixtures_by_form


def resolve_fixture_paths(entries: Sequence[dict[str, object]]) -> list[str]:
    resolved_paths: list[str] = []
    for entry in entries:
        raw_path = entry.get('path')
        if not isinstance(raw_path, str) or not raw_path:
            continue
        path = Path(raw_path)
        full_path = path if path.is_absolute() else ROOT / path
        if full_path.exists():
            resolved_paths.append(str(full_path.resolve()))
    return resolved_paths


def build_row(
    spec: FormSpec,
    workbook_base: Path,
    output_timestamp: str,
    fixtures_by_form: dict[str, list[dict[str, object]]],
) -> tuple[CoverageRow, dict[str, object]]:
    workbook_files = find_matching_files(workbook_base, spec.workbook_patterns)
    workbook_details = [summarize_workbook(path) for path in workbook_files]
    irs_refs = find_matching_files(IRS_MEF_ROOT, spec.irs_reference_patterns)
    local_files = filter_local_files(spec.local_files)
    local_test_files = filter_local_files(spec.local_test_files)
    fixture_entries = fixtures_by_form.get(spec.canonical_form, [])
    fixture_paths = resolve_fixture_paths(fixture_entries)

    if workbook_files and irs_refs:
        coverage_status = 'workbook_and_irs_reference_present'
    elif workbook_files:
        coverage_status = 'workbook_present_without_irs_reference'
    elif irs_refs and fixture_paths:
        coverage_status = 'fixture_backed_irs_reference'
    elif irs_refs:
        coverage_status = 'irs_reference_only'
    else:
        coverage_status = 'missing_external_reference'

    if not local_files:
        coverage_status = 'local_implementation_missing'

    workbook_sheet_names = sorted(
        {sheet for detail in workbook_details for sheet in detail.sheet_names}
    )
    workbook_sheet_count = sum(detail.sheet_count or 0 for detail in workbook_details)

    row = CoverageRow(
        canonical_form=spec.canonical_form,
        workbook_present=bool(workbook_files),
        workbook_paths=[str(path) for path in workbook_files],
        workbook_sheet_names=workbook_sheet_names,
        workbook_sheet_count=workbook_sheet_count,
        irs_reference_present=bool(irs_refs),
        irs_reference_count=len(irs_refs),
        irs_reference_sample_paths=[str(path) for path in irs_refs[:10]],
        local_files=local_files,
        local_file_count=len(local_files),
        local_present=bool(local_files),
        local_test_files=local_test_files,
        local_test_count=len(local_test_files),
        local_tests_present=bool(local_test_files),
        fixture_backed=bool(fixture_paths),
        fixture_count=len(fixture_paths),
        fixture_paths=fixture_paths,
        coverage_status=coverage_status,
        notes=spec.notes,
        last_audited=output_timestamp,
    )

    inventory = {
        'canonical_form': spec.canonical_form,
        'workbook_details': [asdict(detail) for detail in workbook_details],
        'irs_reference_paths': [str(path) for path in irs_refs],
        'local_files': local_files,
        'local_test_files': local_test_files,
        'fixture_entries': fixture_entries,
        'fixture_paths': fixture_paths,
        'notes': spec.notes,
    }
    return row, inventory


def write_csv(rows: Sequence[CoverageRow], path: Path) -> None:
    with path.open('w', newline='') as handle:
        writer = csv.DictWriter(handle, fieldnames=list(asdict(rows[0]).keys()))
        writer.writeheader()
        for row in rows:
            writer.writerow(
                {
                    key: json.dumps(value) if isinstance(value, list) else value
                    for key, value in asdict(row).items()
                }
            )


def write_json(payload: object, path: Path) -> None:
    path.write_text(json.dumps(payload, indent=2, sort_keys=False) + '\n')


def write_markdown(rows: Sequence[CoverageRow], path: Path) -> None:
    lines = [
        '# Business Workbook Parity Harness',
        '',
        'This inventory extends the workbook parity approach beyond the 1040 family.',
        'When a private workbook is unavailable locally, the matrix records the workbook gap',
        'and inventories local IRS MeF/ATS materials instead so parity work stays evidence-based.',
        'Canonical JSON parity fixtures now back 1120-S, 1065, and 1041 so those forms are no longer IRS-reference-led only.',
        '',
        '| Form | Workbook | IRS refs | Fixtures | Local impl | Tests | Status | Notes |',
        '| --- | --- | --- | --- | --- | --- | --- | --- |',
    ]
    for row in rows:
        lines.append(
            f"| {row.canonical_form} | {'yes' if row.workbook_present else 'no'} | {row.irs_reference_count} | {row.fixture_count} | {row.local_file_count} | {row.local_test_count} | {row.coverage_status} | {row.notes} |"
        )
    lines.extend(
        [
            '',
            '## Reading the status',
            '',
            '- `workbook_and_irs_reference_present`: local workbook and IRS support materials are both available.',
            '- `workbook_present_without_irs_reference`: local workbook found, but no IRS support assets were discovered in the scanned materials.',
            '- `fixture_backed_irs_reference`: no local workbook was found, but canonical parity fixtures and IRS materials are both present.',
            '- `irs_reference_only`: no local workbook found, but IRS MeF/ATS assets are available and should anchor parity work until a workbook is supplied.',
            '- `missing_external_reference`: neither a local workbook nor IRS materials were detected for that form.',
            '- `local_implementation_missing`: the form did not resolve to local implementation files, so parity should stop until source support exists.',
            '',
            '## Canonical business fixtures',
            '',
            f'- Fixture manifest: `{FIXTURE_MANIFEST}`',
            '- Fixture-backed forms in this pass: `1120-S`, `1065`, `1041`.',
            '',
        ]
    )
    path.write_text('\n'.join(lines) + '\n')


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        '--workbook-base',
        type=Path,
        default=DOWNLOADS,
        help='Base directory to scan for local business-form workbooks.',
    )
    parser.add_argument(
        '--output-dir',
        type=Path,
        default=DEFAULT_OUTPUT_DIR,
        help='Directory to write docs outputs.',
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    args.output_dir.mkdir(parents=True, exist_ok=True)
    audited_at = datetime.now(timezone.utc).isoformat()
    fixtures_by_form = load_fixture_manifest()

    rows: list[CoverageRow] = []
    inventory: list[dict[str, object]] = []
    for spec in FORM_SPECS:
        row, details = build_row(spec, args.workbook_base, audited_at, fixtures_by_form)
        rows.append(row)
        inventory.append(details)

    write_csv(rows, args.output_dir / 'business_workbook_parity_matrix.csv')
    write_json(
        [asdict(row) for row in rows],
        args.output_dir / 'business_workbook_parity_matrix.json',
    )
    write_json(
        inventory, args.output_dir / 'business_workbook_parity_inventory.json'
    )
    write_markdown(rows, args.output_dir / 'business_workbook_parity_harness.md')
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
